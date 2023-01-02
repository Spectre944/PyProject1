# This Python file uses the following encoding: utf-8

import os
from pathlib import Path
import sys

# import openpyxl module for work with EXEL
import openpyxl

from PySide6.QtGui import QGuiApplication
from PySide6.QtQml import QQmlApplicationEngine
from PySide6.QtCore import QObject, Slot, Signal



class MainWindow(QObject):
    def __init__(self):
        QObject.__init__(self)
        self.openWorkBook("vocab.xlsx")

    #Signal to QML
    wordProcessed = Signal(str)

    #slot from QML
    @Slot(str)
    def getWord(self, word):
        if len(word):
            #ws.write(0, 0, word, style0)
            #wb.save('Vocab.xlsx')
            self.wordProcessed.emit(word)
        else:
            self.wordProcessed.emit("Nothing to write")


    def openWorkBook(self, pathStr):
        if os.path.exists(pathStr):
            # To open the workbook
            # workbook object is created
            wb = openpyxl.load_workbook(pathStr)
            print("WorkBook opened")

            #open active sheet of WorkBook
            sheet = wb.active

            # Getting the value of maximum rows
            # and column
            row = sheet.max_row
            column = sheet.max_column

#            for i in range(1, row + 1):
#                cell_obj = sheet.cell(row = i, column = 3)
#                print(cell_obj.value)

            print("Total Rows:", row)
            print("Total Columns:", column)


        else:
            wb = openpyxl.Workbook()
            wb.save(filename = "vocab.xlsx")
            print("WorkBook not exists\n"
                  "Creating new one")





if __name__ == "__main__":
    app = QGuiApplication(sys.argv)
    engine = QQmlApplicationEngine()

    main = MainWindow()
    engine.rootContext().setContextProperty("backend", main)


    engine.load(os.fspath(Path(__file__).resolve().parent / "main.qml"))
    if not engine.rootObjects():
        sys.exit(-1)
    sys.exit(app.exec())
